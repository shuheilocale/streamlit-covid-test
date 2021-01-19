import pandas as pd
import openpyxl

def read_patients_xlsx(fname):
    # 療養状況等及び入院患者受入病床数等に関する調査について
    # https://www.mhlw.go.jp/stf/seisakunitsuite/newpage_00023.html
    # ここのエクセルファイルを読み込む
    wb = openpyxl.load_workbook(fname)
    ws = wb.worksheets[0]
    area = ws['C8:C54']
    n_patients = ws['D8:D54']
    n_patients_hospital = ws['E8:E54']
    capacity_hospital = ws['G8:G54']
    n_patients_inn = ws['O8:O54']
    capacity_inn = ws['Q8:Q54']

    data = []

    for area, n_p, n_p_hos, cap_hos, n_p_inn, cap_inn in \
        zip(area, n_patients, n_patients_hospital, capacity_hospital, n_patients_inn, capacity_inn):

        data.append(
            [area[0].value.split(' ')[0], n_p[0].value, n_p_hos[0].value, cap_hos[0].value, n_p_inn[0].value, cap_inn[0].value])

    df = pd.DataFrame(data, columns=['id', 'n_patients', 'n_patients_hospital', 'capacity_hospital', 'n_patients_inn', 'capacity_inn'])
    df['id'] = df['id'].astype(int)

    return df

if __name__ == '__main__':
    df = read_patients_xlsx('000721104.xlsx')
    print(df)